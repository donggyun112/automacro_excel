import signal
import sys
from tkinter import filedialog
import openpyxl.utils
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import unicodedata
from tabulate import tabulate

image_cells = [()]

def check_image_in_cell(file_name, cell):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    
    if cell.value is None and cell.has_style:
        for obj in sheet._images:
			  # 수정된 부분
            # cell_col = cell.column
            # cell_row = cell.row
            # print(obj.anchor._from.col, obj.anchor._from.row,  cell.row, cell.column)
            if (obj.anchor._from.row + 1 == cell.row) and (obj.anchor._from.col + 1 == cell.column):  # 수정된 부분
                return True
            # if obj.anchor._from.col == cell.coordinate:  # 수정된 부분
                # print(obj.anchor._from)
    
    # print("test : false")
    return False

def signal_handler(sig, frame):
    print("=" * 50)
    print("{:^50}".format("프로그램을 종료합니다."))
    print("=" * 50)
    sys.exit(0)



# row_index -> 행 인덱스
# col_idx -> 열 인덱스
def find_sheet_name(target_value):
    global tf
    idx = 1
    row_idx = 0
    tmp = tf
    max_rows = tmp.shape[0]
    
    
    for row_idx in range(max_rows):
        try:
            for column_name in target_value:
                if column_name in tmp.columns:
                    print("=" * 50)
                    print("{:^50}".format(f"'{row_idx + 1}'번째 행에 '{column_name}' 열이 존재합니다"))
                    print("=" * 50)
                    return row_idx + 1, tmp.columns.get_loc(column_name[0]), tmp
        except ValueError:
            continue
        print(tmp.columns)
        tmp = tmp.iloc[1:].reset_index(drop=True)
        tmp.columns = tmp.iloc[0]
        tmp = tmp[1:] 
    print("=" * 50)
    print("{:^50}".format("유효한 헤더 행을 찾을 수 없습니다."))
    print("=" * 50)
    return -1, -1



def preformat_cell(cell, width, align='<', fill=' '):
    count = (width - sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in str(cell)))
    return {
        '>': lambda s: fill * count + s,
        '<': lambda s: s + fill * count,
        '^': lambda s: fill * (count // 2) + s + fill * (count // 2 + count % 2)
    }[align](str(cell))

def show_table(col_idx, image_column):
    global image_cells
    try:
        headers = df.columns.tolist()
        data = df.values.tolist()

        if image_column == 'show':
            selected_column_index = 0
        else:
            selected_column_index = headers.index(image_column)

        formatted_data = []
        for row_idx, row in enumerate(data, start=col_idx+2):
            formatted_row = []
            for col_idx, cell in enumerate(row, start=1):
                if (row_idx, col_idx) in image_cells:
                    formatted_cell = f"\033[41m{str(cell)}\033[0m"
                elif col_idx == selected_column_index + 1 and image_column != 'show':
                    formatted_cell = f"\033[103m{str(cell)}\033[0m"
                else:
                    formatted_cell = str(cell)
                formatted_row.append(formatted_cell)
            formatted_data.append(formatted_row)

        formatted_headers = [f"\033[103m{header}\033[0m" if header == image_column else header for header in headers]
        table = tabulate(formatted_data, headers=formatted_headers, tablefmt="grid")

        print("=" * 50)
        print(table)
        print("=" * 50)
        return table
    except Exception as e:
        print("=" * 50)
        print("{:^50}".format(f"테이블을 출력하는 중 오류가 발생했습니다: {str(e)}"))
        print("=" * 50)


def insert_images(file_name, image_column):
    global image_cells
    column_index = df.columns.get_loc(image_column)
    column_letter = openpyxl.utils.get_column_letter(column_index + 1)

    workbook = load_workbook(file_name)
    sheet = workbook.active
    column_width = sheet.column_dimensions[column_letter].width

    row_heights = []
    for i, cell in enumerate(sheet[column_letter], start=1):
        if i >= col_idx:
            row_heights.append(sheet.row_dimensions[cell.row].height)

    try:
        for i, row in df.iterrows():
            cell = sheet.cell(row=i + (col_idx), column=column_index + 1)
            if check_image_in_cell(file_name, cell) or cell.value:
                print("{:^50}".format(f"이미지가 이미 있습니다. 행: {i + column_index + 1}"))
                continue
            else:
                print("{:^50}".format(f">> 이미지를 삽입합니다. 행: {i + column_index + 1}"))

                file_path = filedialog.askopenfilename(parent=None, filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif")])

                if file_path:
                    img = Image(file_path)

                    img_width = column_width * 7.8
                    img_height = row_heights[i] * 1.28
                    img.width = img_width
                    img.height = img_height

                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    sheet.add_image(img, cell.coordinate)
                    # cell.value = " "
                    image_cells.append((i + col_idx, column_index + 1))
                else:
                    print("{:^50}".format(">> 작업을 취소합니다."))
                    break
    except Exception as e:
        print("=" * 50)
        print("{:^50}".format(f"이미지 삽입 중 오류가 발생했습니다: {str(e)}"))
        print("{:^50}".format("프로그램을 종료합니다."))
        print("=" * 50)
        exit()

    print("=" * 50)
    print("{:^50}".format("현재 작업 내용을 저장합니다."))
    print("=" * 50)
    output_file = 'output.xlsx'
    workbook.save(output_file)
    return output_file


def find_image_cells(file_name):
    global image_cells
    workbook = load_workbook(file_name)
    sheet = workbook.active
    for obj in sheet._images:
        image_cells.append((obj.anchor._from.row, obj.anchor._from.col))

def insert_images_history(file_name, image_column):
    
    global image_cells
    column_index = df.columns.get_loc(image_column)
    workbook = load_workbook(file_name)
    sheet = workbook.active
    for i, row in df.iterrows():
        cell = sheet.cell(row=i + (col_idx), column=column_index + 1)
        if check_image_in_cell(file_name, cell):
            image_cells.append((i + col_idx, column_index + 1))


# def set_table():
#     global tf, df
#     df = tf.iloc[row_idx:]
#     df.columns = df.iloc[0]
#     df = df[1:]
	

if __name__ == "__main__":
    signal.signal(signal.SIGINT, signal_handler)
    print("=" * 50)
    print("{:^50}".format("Excel 파일을 선택하세요."))
    print("=" * 50)
    file_name = filedialog.askopenfilename(parent=None, filetypes=[("Excel files", "*.xlsx")])

    if not file_name:
        print("=" * 50)
        print("{:^50}".format("파일을 선택하지 않았습니다."))
        print("{:^50}".format("프로그램을 종료합니다."))
        print("=" * 50)
        exit()
    # def init_history_image():
        
    # init_history_image()
    tf = pd.read_excel(file_name, header=None, engine='openpyxl')
    # find_image_cells(file_name)
    while True:
        while True:
            print("-" * 50)
            image_column = input("열 이름을 입력하세요: ")
            input_column = image_column
            image_column = image_column.split(" ")
            print(image_column, type(image_column))
            if input_column == "show table" or input_column == "테이블 보기" or input_column == "테이블":
                df = tf
                find_image_cells(file_name)
                show_table(0, 'show')
                continue
            try:
                row_idx, col_idx, df = find_sheet_name(image_column)
                if col_idx == -1:
                    print("{:^50}".format("다시 입력해 주세요."))
                    continue
                # set_table()
                insert_images_history(file_name, image_column[0])
                table = show_table(col_idx, image_column[0])
                break
            except Exception as e:
                print("=" * 50)
                print("{:^50}".format(f"파일을 읽는 중 오류가 발생했습니다: {str(e)}"))
                print("{:^50}".format("올바른 Excel 파일을 선택해 주세요."))
                print("=" * 50)

        image_column = image_column[0]
        output_file = insert_images(file_name, image_column)

        while True:
            try:
                print("=" * 50)
                answer = input("작업을 계속하시겠습니까? (y/n): ")
                if answer == "show table" or answer == "테이블 보기" or answer== "테이블":
                    show_table(0, 'show')
                    continue
                if answer.lower() in ['y', 'yes', '네', 'ㅛ']:
                    file_name = output_file
                    break
                elif answer.lower() in ['n', 'no', '아니오', 'ㅜ']:
                    print("=" * 50)
                    print("{:^50}".format("프로그램을 종료합니다."))
                    print("=" * 50)
                    exit()
                else:
                    print("=" * 50)
                    print("{:^50}".format("잘못된 입력입니다. 다시 입력해 주세요."))
                    print("=" * 50)
            except Exception as e:
                print("=" * 50)
                print("{:^50}".format(f"입력 중 오류가 발생했습니다: {str(e)}"))
                print("{:^50}".format("다시 입력해 주세요."))
                print("=" * 50)
                continue