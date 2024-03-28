from openpyxl import load_workbook

def check_image(file_path, cell_address):
    wb = load_workbook(file_path)
    ws = wb.active
    print(ws['D5'])
    exit()
    cell = ws[cell_address]
    if cell.has_hyperlink:
        if 'image' in cell.hyperlink.target:
            print(f"셀 {cell_address}에 이미지가 있습니다.")
            return True
    
    print(f"셀 {cell_address}에 이미지가 없습니다.")
    return False

# 사용 예시
file_path = "output.xlsx"
cell_address = "사진"
check_image(file_path, cell_address)