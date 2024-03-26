import os
import signal
import sys
from tkinter import filedialog
from tkintertable import TableCanvas

import openpyxl.utils
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import unicodedata
from wcwidth import wcswidth
from tabulate import tabulate
from openpyxl.styles import PatternFill
import tkinter as tk


def signal_handler(sig, frame):
    print("=" * 50)
    print("{:^50}".format("프로그램을 종료합니다."))
    print("=" * 50)
    sys.exit(0)


def find_sheet_name(file_name, column_names):
    try:
        df = pd.read_excel(file_name, header=None)
        max_rows = df.shape[0]

        for i in range(max_rows):
            try:
                df = pd.read_excel(file_name, header=i)

                for column_name in column_names:
                    if column_name in df.columns:
                        print("=" * 50)
                        print("{:^50}".format(f"'{i}'번째 행에 '{column_name}' 열이 존재합니다"))
                        print("=" * 50)
                        return i
            except ValueError:
                continue

        print("=" * 50)
        print("{:^50}".format("유효한 헤더 행을 찾을 수 없습니다."))
        print("=" * 50)
        return -1
    except FileNotFoundError:
        print("=" * 50)
        print("{:^50}".format(f"'{file_name}' 파일을 찾을 수 없습니다."))
        print("=" * 50)
        return -1

def preformat_cell(cell, width, align='<', fill=' '):
    count = (width - sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in str(cell)))
    return {
        '>': lambda s: fill * count + s,
        '<': lambda s: s + fill * count,
        '^': lambda s: fill * (count // 2) + s + fill * (count // 2 + count % 2)
    }[align](str(cell))

def show_table(file_name, col_idx, image_column):
    try:
        df = pd.read_excel(file_name, header=col_idx)
        headers = df.columns.tolist()
        data = df.values.tolist()

        # 선택한 열의 인덱스 찾기
        selected_column_index = headers.index(image_column)

        # 선택한 열에 하이라이팅 효과 적용
        formatted_data = []
        for row in data:
            formatted_row = [f"\033[103m{str(cell)}\033[0m" if idx == selected_column_index else str(cell) for idx, cell in enumerate(row)]
            formatted_data.append(formatted_row)

        formatted_headers = [f"\033[103m{header}\033[0m" if header == image_column else header for header in headers]

        table = tabulate(formatted_data, headers=formatted_headers, tablefmt="grid")
        print("=" * 50)
        print(table)
        print("=" * 50)
    except Exception as e:
        print("=" * 50)
        print("{:^50}".format(f"테이블을 출력하는 중 오류가 발생했습니다: {str(e)}"))
        print("=" * 50)

import tkinter as tk
from tkintertable import TableCanvas
import pandas as pd
import unicodedata

def show_table_gui(file_name, col_idx, image_column):
    try:
        df = pd.read_excel(file_name, header=col_idx)
        headers = df.columns.tolist()
        data = df.astype(str).values.tolist()  # 데이터를 문자열로 변환

        # 선택한 열의 인덱스 찾기
        selected_column_index = headers.index(image_column)
        selected_column_index : int
        print("ok")

        # tkinter 창 생성
        root = tk.Tk()
        root.title("Table Viewer")

        # tkintertable 테이블 생성
        table = TableCanvas(root, data=data, headers=headers)

        print("ok")

        table.show()

        print("ok")

        # 선택한 열에 하이라이팅 효과 적용
        for row_idx in range(len(data)):
            row_color = 'yellow' if row_idx % 2 == 0 else 'white'
            table.setRowColors(row=row_idx, rowColor=row_color)
            table.setCellColor(row=row_idx, col=headers[selected_column_index], color='lightblue')

        # 셀 너비 조정
        col_widths = [max(len(str(cell)) + 2 for cell in col) for col in zip(*data)]
        for col_idx, width in enumerate(col_widths):
            table.setColumnWidth(col_idx, width)

        # GUI 이벤트 루프 시작
        root.mainloop()

    except Exception as e:
        print("=" * 50)
        print("{:^50}".format(f"테이블을 출력하는 중 오류가 발생했습니다: {str(e)}"))
        print("=" * 50)


def insert_images(file_name, image_column):
    column_index = df.columns.get_loc(image_column)
    column_letter = openpyxl.utils.get_column_letter(column_index + 1)

    workbook = load_workbook(file_name)
    sheet = workbook.active
    column_width = sheet.column_dimensions[column_letter].width

    row_heights = []
    for i, cell in enumerate(sheet[column_letter], start=1):
        if i >= col_idx + 2:
            row_heights.append(sheet.row_dimensions[cell.row].height)

    try:
        for i, row in df.iterrows():
            cell = sheet.cell(row=i + (col_idx + 2), column=column_index + 1)

            if cell.value:
                print("{:^50}".format(f"이미지가 이미 있습니다. 행: {i + column_index + 1}"))
                continue
            else:
                print("{:^50}".format(f">> 이미지를 삽입합니다. 행: {i + column_index + 1}"))

                file_path = filedialog.askopenfilename(parent=None, filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif")])

                if file_path:
                    img = Image(file_path)

                    img_width = column_width * 7.8
                    img_height = row_heights[i] * 1.28
                    img.width = img_width
                    img.height = img_height

                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    sheet.add_image(img, cell.coordinate)
                    cell.value = " "
                else:
                    print("{:^50}".format(">> 작업을 취소합니다."))
                    break
    except Exception as e:
        print("=" * 50)
        print("{:^50}".format(f"이미지 삽입 중 오류가 발생했습니다: {str(e)}"))
        print("{:^50}".format("프로그램을 종료합니다."))
        print("=" * 50)
        exit()

    print("=" * 50)
    print("{:^50}".format("현재 작업 내용을 저장합니다."))
    print("=" * 50)
    output_file = 'output.xlsx'
    workbook.save(output_file)
    return output_file


if __name__ == "__main__":
    signal.signal(signal.SIGINT, signal_handler)
    print("=" * 50)
    print("{:^50}".format("Excel 파일을 선택하세요."))
    print("=" * 50)
    file_name = filedialog.askopenfilename(parent=None, filetypes=[("Excel files", "*.xlsx")])

    if not file_name:
        print("=" * 50)
        print("{:^50}".format("파일을 선택하지 않았습니다."))
        print("{:^50}".format("프로그램을 종료합니다."))
        print("=" * 50)
        exit()
    while True:
        while True:
            print("-" * 50)
            image_column = input("열 이름을 입력하세요: ").split()

            try:
                col_idx = find_sheet_name(file_name, image_column)
                if col_idx == -1:
                    print("{:^50}".format("다시 입력해 주세요."))
                    continue
                df = pd.read_excel(file_name, header=col_idx)
                #show_table(file_name, col_idx, image_column[0])
                show_table_gui(file_name, col_idx, image_column[0])
                break
            except Exception as e:
                print("=" * 50)
                print("{:^50}".format(f"파일을 읽는 중 오류가 발생했습니다: {str(e)}"))
                print("{:^50}".format("올바른 Excel 파일을 선택해 주세요."))
                print("=" * 50)
        image_column = image_column[0]
        output_file = insert_images(file_name, image_column)

        while True:
            try:
                print("=" * 50)
                answer = input("작업을 계속하시겠습니까? (y/n): ")
                if answer.lower() in ['y', 'yes', '네', 'ㅛ']:
                    file_name = output_file
                    break
                elif answer.lower() in ['n', 'no', '아니오', 'ㅜ']:
                    print("=" * 50)
                    print("{:^50}".format("프로그램을 종료합니다."))
                    print("=" * 50)
                    exit()
                else:
                    print("=" * 50)
                    print("{:^50}".format("잘못된 입력입니다. 다시 입력해 주세요."))
                    print("=" * 50)
            except Exception as e:
                print("=" * 50)
                print("{:^50}".format(f"입력 중 오류가 발생했습니다: {str(e)}"))
                print("{:^50}".format("다시 입력해 주세요."))
                print("=" * 50)
                continue