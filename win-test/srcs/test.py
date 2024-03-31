import signal
import sys
import os
from tkinter import filedialog
import openpyxl.utils
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import unicodedata
from tabulate import tabulate

image_cells = []
workbook = None
sheet = None
selected_sheet = None
tf = None
flag = True

def select_sheet(file_name):
    global workbook
    try:
        workbook = load_workbook(file_name, data_only=True)
        sheets = workbook.sheetnames
        print("=" * 50)
        print("{:^50}".format("시트 목록"))
        print("=" * 50)
        for i, sheet_name in enumerate(sheets, start=1):
            print(f"{sheet_name}")
        print("=" * 50)
        while True:
            try:
                sheet_num = int(input("사용할 시트의 번호를 입력하세요: "))
                if 1 <= sheet_num <= len(sheets):
                    selected_sheet = sheets[sheet_num - 1]
                    print("=" * 50)
                    print("{:^50}".format(f"선택한 시트: {selected_sheet}"))
                    print("=" * 50)
                    return selected_sheet
                else:
                    print("=" * 50)
                    print("{:^50}".format("유효하지 않은 시트 번호입니다. 다시 입력해 주세요."))
                    print("=" * 50)
            except ValueError:
                print("=" * 50)
                print("{:^50}".format("숫자를 입력해 주세요."))
                print("=" * 50)
    except Exception as e:
        print("=" * 50)
        print("{:^50}".format("파일을 읽는 중 오류가 발생했습니다."))
        print("{:^50}".format(str(e)))
        print("=" * 50)
        return None

# def check_image_in_cell(cell):
#     global workbook, sheet
#     if cell.value is None and cell.has_style:
#         for obj in sheet._images:
#             if (obj.anchor._from.row + 1 == cell.row) and (obj.anchor._from.col + 1 == cell.column):
#                 return True
#     return False

def check_image_in_cell(cell):
    global image_cells
    if not image_cells or image_cells == []:
        return False
    for row, col in image_cells:
        if cell.row == row and cell.column == col:
            return True
    return False



def signal_handler(sig, frame):
    print("=" * 50)
    print("{:^50}".format("프로그램을 종료합니다."))
    print("=" * 50)
    sys.exit(0)

def find_sheet_name(column_names):
    try:
        max_rows = tf.shape[0]
        for i in range(max_rows):
            try:
                tmp = pd.read_excel(file_name, sheet_name=selected_sheet, header=i)
                for column_name in column_names:
                    if column_name in tmp.columns:
                        print("=" * 50)
                        print("{:^50}".format(f"'{i}'번째 행에 '{column_name}' 열이 존재합니다"))
                        print("=" * 50)
                        return i, tmp
                # print(tmp.columns)
            except ValueError:
                continue
        print("=" * 50)
        print("{:^50}".format("유효한 헤더 행을 찾을 수 없습니다."))
        print("=" * 50)
        return -1, None
    except FileNotFoundError:
        print("=" * 50)
        print("{:^50}".format(f"'{file_name}' 파일을 찾을 수 없습니다."))
        print("=" * 50)
        return -1, None

def preformat_cell(cell, width, align='<', fill=' '):
    count = (width - sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in str(cell)))
    return {
        '>': lambda s: fill * count + s,
        '<': lambda s: s + fill * count,
        '^': lambda s: fill * (count // 2) + s + fill * (count // 2 + count % 2)
    }[align](str(cell))

def show_table(df, col_idx, image_column):
    global image_cells, sheet
    try:
        headers = df.columns.tolist()
        data = df.values.tolist()
        if image_column == 'show':
            selected_column_index = 0
        else:
            selected_column_index = headers.index(image_column)
        formatted_data = []
        for row_idx, row in enumerate(data, start=col_idx+2):
            formatted_row = []
            for col_idx, cell in enumerate(row, start=1):
                cell_obj = sheet.cell(row=row_idx, column=col_idx)
                cell_value = cell
                for merged_range in sheet.merged_cells.ranges:
                    if cell_obj.coordinate in merged_range:
                        # 병합된 셀인 경우, 병합된 범위의 첫 번째 셀 값을 사용
                        cell_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                        break
                # 셀의 값이 수식인 경우, 수식의 결과 값을 가져옴
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if (row_idx, col_idx) in image_cells:
                    formatted_cell = f"\033[41m{str(cell_value)}\033[0m"
                elif col_idx == selected_column_index + 1 and image_column != 'show':
                    formatted_cell = f"\033[103m{str(cell_value)}\033[0m"
                else:
                    formatted_cell = str(cell_value)
                formatted_row.append(formatted_cell)
            formatted_data.append(formatted_row)
        formatted_headers = [f"\033[103m{header}\033[0m" if header == image_column else header for header in headers]
        table = tabulate(formatted_data, headers=formatted_headers, tablefmt="grid")
        print("=" * 50)
        print(table)
        print("=" * 50)
        return table
    except Exception as e:
        print("=" * 50)
        print("{:^50}".format(f"테이블을 출력하는 중 오류가 발생했습니다: {str(e)}"))
        print("=" * 50)

def insert_images(image_column, output_file):
    global image_cells, flag
    check_merge = False

    workbook = load_workbook(file_name)
    sheet = workbook[selected_sheet]

    column_index = df.columns.get_loc(image_column)
    column_letter = openpyxl.utils.get_column_letter(column_index + 1)

    row_heights = []
    for i, cell in enumerate(sheet[column_letter], start=1):
        if i >= col_idx + 2:
            row_heights.append(sheet.row_dimensions[cell.row].height)

    # 셀이 병합되었는지 확인
    cell = sheet.cell(row=col_idx + 1, column=column_index + 1)
    if cell.coordinate in sheet.merged_cells:
        # 병합된 셀의 범위 찾기
        merged_range = None
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                merged_range = range_
                break
        if merged_range:
            # 병합된 셀의 가로 길이 계산
            print("{:^50}".format(f"병합된 셀: {merged_range}"))
            check_merge = True
            start_column = merged_range.min_col
            end_column = merged_range.max_col
            column_width = 0
            for column in range(start_column, end_column + 1):
                column_letter = openpyxl.utils.get_column_letter(column)
                column_width += sheet.column_dimensions[column_letter].width
        else:
            print("{:^50}".format("병합된 셀이지만, 병합된 범위를 찾을 수 없습니다."))
            return
    else:
        column_width = sheet.column_dimensions[column_letter].width

    try:
        for i, row in df.iterrows():
            cell = sheet.cell(row=i + (col_idx + 2), column=column_index + 1)

            if check_image_in_cell(cell) or cell.value:
                print("{:^50}".format(f"이미지가 이미 있습니다. 행: {i + col_idx + 1}"))
                continue
            else:
                print("{:^50}".format(f">> 이미지를 삽입합니다. 행: {i + col_idx + 1}"))
                file_path = filedialog.askopenfilename(parent=None, filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif")])
                if file_path:
                    img = Image(file_path)
                    img_width = column_width * 7.8
                    img_height = row_heights[i] * 1.28
                    img.width = img_width
                    img.height = img_height
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    sheet.add_image(img, cell.coordinate)
                    if check_merge:
                        for column in range(start_column, end_column + 1):
                            image_cells.append((i + col_idx + 2, column))
                    else:
                        image_cells.append((i + col_idx + 2, column_index + 1))
                else:
                    print("{:^50}".format(">> 작업을 취소합니다."))
                    break
    except Exception as e:
        print("=" * 50)
        print("{:^50}".format(f"이미지 삽입 중 오류가 발생했습니다: {str(e)}"))
        print("{:^50}".format("프로그램을 종료합니다."))
        print("=" * 50)
        exit()
    print("=" * 50)
    print("{:^50}".format("현재 작업 내용을 저장합니다."))
    print("=" * 50)
    if not os.path.exists('output.xlsx'):
        output_file = 'output.xlsx'
    elif flag == True:
        print("=" * 50)
        print("{:^50}".format("이미 'output.xlsx' 파일이 존재합니다."))
        print("=" * 50)
        i = 1
        while flag:
            output_file = f'output_{i}.xlsx'
            if not os.path.exists(output_file):
                break
            else:
                print("=" * 50)
                print("{:^50}".format(f"'{output_file}' 파일이 이미 존재합니다."))
                print("=" * 50)
            i += 1
    print("=" * 50)
    print("{:^50}".format(f"'{output_file}' 파일로 저장합니다."))
    print("=" * 50)
    workbook.save(output_file)
    flag = False
    return output_file

def insert_images_history(image_column):
    global image_cells, sheet
    column_index = df.columns.get_loc(image_column)
    print("=" * 50)
    print(col_idx, column_index)
    for i, row in df.iterrows():
        cell = sheet.cell(row=i + (col_idx + 2), column=column_index + 1)
        if check_image_in_cell(cell) or cell.value:
            image_cells.append((i + col_idx + 2, column_index + 1))
            
def find_image_cells():
    global image_cells, sheet
    for obj in sheet._images:
        image_cells.append((obj.anchor._from.row + 1, obj.anchor._from.col + 1))

if __name__ == "__main__":
    output_file = None
    signal.signal(signal.SIGINT, signal_handler)
    print("=" * 50)
    print("{:^50}".format("Excel 파일을 선택하세요."))
    print("=" * 50)
    file_name = filedialog.askopenfilename(parent=None, filetypes=[("Excel files", "*.xlsx")])
    if not file_name:
        print("=" * 50)
        print("{:^50}".format("파일을 선택하지 않았습니다."))
        print("{:^50}".format("프로그램을 종료합니다."))
        print("=" * 50)
        exit()
    selected_sheet = select_sheet(file_name)
    if selected_sheet is None:
        print("=" * 50)
        print("{:^50}".format("프로그램을 종료합니다."))
        print("=" * 50)
        exit()
    tf = pd.read_excel(file_name, sheet_name=selected_sheet, header=0)
    sheet = workbook[selected_sheet]
    find_image_cells()
    show_table(tf.copy(), 0, 'show')
    while True:
        while True:
            print("-" * 50)
            image_column = input("열 이름을 입력하세요: ")
            input_column = image_column
            image_column = image_column.split(" ")
            print(image_column, type(image_column))
            if input_column == "show table" or input_column == "테이블 보기" or input_column == "테이블":
                show_table(tf.copy(), 0, 'show')
                continue
            try:
                col_idx, df = find_sheet_name(image_column)
                if col_idx == -1:
                    print("{:^50}".format("다시 입력해 주세요."))
                    continue
                insert_images_history(image_column[0])
                table = show_table(df.copy(), col_idx, image_column[0])
                break
            except Exception as e:
                print("=" * 50)
                print("{:^50}".format(f"파일을 읽는 중 오류가 발생했습니다: {str(e)}"))
                print("{:^50}".format("올바른 Excel 파일을 선택해 주세요."))
                print("=" * 50)
        image_column = image_column[0]
        output_file = insert_images(image_column, output_file)
        while True:
            try:
                print("=" * 50)
                answer = input("작업을 계속하시겠습니까? (y/n): ")
                if answer == "show table" or answer == "테이블 보기" or answer== "테이블":
                    show_table(tf.copy(), 0, 'show')
                    continue
                if answer.lower() in ['y', 'yes', '네', 'ㅛ', "sp"]:
                    file_name = output_file
                    break
                elif answer.lower() in ['n', 'no', '아니오', 'ㅜ', 'dksldh']:
                    print("=" * 50)
                    print("{:^50}".format("프로그램을 종료합니다."))
                    print("=" * 50)
                    exit()
                else:
                    print("=" * 50)
                    print("{:^50}".format("잘못된 입력입니다. 다시 입력해 주세요."))
                    print("=" * 50)
            except Exception as e:
                print("=" * 50)
                print("{:^50}".format(f"입력 중 오류가 발생했습니다: {str(e)}"))
                print("{:^50}".format("다시 입력해 주세요."))
                print("=" * 50)
                continue