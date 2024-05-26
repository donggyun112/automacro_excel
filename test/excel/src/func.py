import pandas as pd
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.cell_range import CellRange
from PyQt5.QtWidgets import QMessageBox

# prev_month_keyword: 지켜야져야하는 열 current_month_keyword: curr -> prev 로 업데이트할 열
# 업데이트할 열이 없으면 NameError 반환

def update_prev_month_col(workbook, sheet_name, previous_month_keyword, current_month_keyword):
    if not previous_month_keyword or not current_month_keyword:
        return NameError("Keyword not found")
    worksheet = workbook[sheet_name]
    strip_keyword = lambda x: x.replace(' ', '').replace('\n', '').replace('\t', '')
    previous_month_keyword = strip_keyword(previous_month_keyword)
    current_month_keyword = strip_keyword(current_month_keyword)
    if previous_month_keyword == current_month_keyword:
        return NameError("키워드가 같음")
    elif previous_month_keyword == '' or current_month_keyword == '':
        return NameError("키워드가 비어있음")
    
    prev_month_col = None
    curr_month_col = None

    # 모든 행을 탐색하며 '전월', '당월' 셀 찾기
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            tmp_cell_value = strip_keyword(str(cell_value))
            if tmp_cell_value == previous_month_keyword:
                print(f"prev_month_col: {col}")
                prev_month_col = col
            elif tmp_cell_value == current_month_keyword:
                print(f"curr_month_col: {col}")
                curr_month_col = col

            # '전월'과 '당월' 열이 모두 찾아졌으면 데이터 덮어쓰기 작업 수행
            if prev_month_col and curr_month_col:
                for data_row in range(row+1, worksheet.max_row + 1):
                    prev_cell = worksheet.cell(row=data_row, column=prev_month_col)
                    curr_cell = worksheet.cell(row=data_row, column=curr_month_col)
                    
                    if isinstance(prev_cell, openpyxl.cell.MergedCell):
                        continue  # 병합된 셀은 건너뛰기
                    elif prev_cell.data_type == 's':  # 문자열인 경우
                        prev_cell.value = curr_cell.value
                    elif prev_cell.data_type == 'f':  # 함수나 수식인 경우
                        if curr_cell.data_type == 'f':  # 수식인 경우
                            continue
                            # prev_cell.value = '=' + str(Translator(curr_cell.value, origin='A1').translate_formula('A1'))
                        else:
                            continue  # 함수인 경우
                            prev_cell.value = curr_cell.value
                            prev_cell.data_type = 'n'  # 데이터 타입을 숫자로 설정
                            prev_cell.value = '=' + str(Translator(prev_cell.value, origin='A1').translate_formula('A1'))
                    else:  # 그 외의 경우 (숫자, 날짜 등)
                        prev_cell.value = curr_cell.value
                        
                print(f"'{sheet_name}' 시트의 {row}번째 행에서 '{current_month_keyword}' 열의 값이 '{previous_month_keyword}' 열로 업데이트되었습니다.")
                return  None# 작업 완료 후 함수 종료

    print(f"'{sheet_name}' 시트에서 '전월'과 '당월' 열을 찾을 수 없습니다.")
    return NameError("키워드를 찾을 수 없음")

def merge_sheets(wb1, wb2, sheet_name=0, range_: CellRange = None, _range_="All", prev_month_keyword=None, curr_month_keyword=None):
    statuses = True
    status = None
    if _range_ == "All" and prev_month_keyword and curr_month_keyword:
        status = update_prev_month_col(wb1, sheet_name, prev_month_keyword, curr_month_keyword)
    try:
        strip_keyword = lambda x: x.replace(' ', '').replace('\n', '').replace('\t', '')
        if prev_month_keyword and curr_month_keyword:
            prev_month_keyword = strip_keyword(prev_month_keyword)
            curr_month_keyword = strip_keyword(curr_month_keyword)
        # 시트 이름으로 시트 가져오기
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        # 병합 범위 설정
        if range_ is None:
            merge_range = CellRange(min_col=1, min_row=1, max_col=sheet1.max_column, max_row=sheet1.max_row)
        else:
            merge_range = range_

        # 병합 범위의 데이터를 리스트로 변환
        data2 = [[cell.value for cell in row] for row in sheet2.iter_rows(min_row=merge_range.min_row, max_row=merge_range.max_row, min_col=merge_range.min_col, max_col=merge_range.max_col)]

        # 리스트를 DataFrame으로 변환
        df2 = pd.DataFrame(data2)

        # 병합된 데이터를 wb1의 시트에 저장
        for col in range(merge_range.min_col, merge_range.max_col+1):
            has_keyword = False
            for row in range(merge_range.min_row, merge_range.max_row+1):
                cell = sheet1.cell(row=row, column=col)
                if isinstance(cell, openpyxl.cell.MergedCell):
                    continue # 병합된 셀은 건너뛰기
                elif prev_month_keyword and strip_keyword(str(cell.value)).find(prev_month_keyword) != -1:
                    has_keyword = True
                    break

            if has_keyword:
                continue # '전월' 키워드가 있는 열은 건너뛰기

            for row in range(merge_range.min_row, merge_range.max_row+1):
                cell = sheet1.cell(row=row, column=col)
                if isinstance(cell, openpyxl.cell.MergedCell):
                    continue # 병합된 셀은 건너뛰기
                elif cell.data_type == 'f': # 셀에 수식이나 함수가 있는 경우
                    formula_value = '=' + str(Translator(cell.value, origin='A1').translate_formula('A1'))
                    cell.value = df2.at[row-merge_range.min_row, col-merge_range.min_col]
                    cell.data_type = 'f'
                    cell.value = formula_value
                else:
                    cell.value = df2.at[row-merge_range.min_row, col-merge_range.min_col]

    except Exception as e:
        print(f"오류 발생: {str(e)}")
        statuses = False
    finally:
        wb1.close()
        wb2.close()
        return statuses, status